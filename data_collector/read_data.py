import openpyxl
import re
data = openpyxl.load_workbook("data\datasheet.xlsx", data_only = True)


def eachMenus(worksheet):
    menus = []
    res = 1
    for week in range(5):
        res += 1
        val = worksheet.cell(row=res, column=1).value
        while val != '날짜':
            if val == '아침': border1 = res
            if val == '점심': border2 = res
            if val == '저녁': border3 = res
            res += 1
            val = worksheet.cell(row=res, column=1).value
        border4 = res

        for day in range(2,9):
            breakfast, lunch, dinner = [], [], []
            flag = 0
            for row in range(border1, border2):
                tmpmenu = worksheet.cell(row=row, column=day).value
                if tmpmenu:
                    splited_menu = re.compile('[가-힣]+').findall(tmpmenu)
                    if not splited_menu: continue
                    menu = splited_menu[0]
                    for k in range(1,len(splited_menu)):
                        menu += '/'
                        menu += splited_menu[k]
                    if tmpmenu[0] == '/' or flag == 1:
                        last = breakfast.pop()
                        breakfast.append(last + '/' + menu)
                        flag = 0
                    else: breakfast.append(menu)
                    if tmpmenu[-1] == '/':
                        flag = 1
            for row in range(border2, border3):
                tmpmenu = worksheet.cell(row=row, column=day).value
                if tmpmenu:
                    splited_menu = re.compile('[가-힣]+').findall(tmpmenu)
                    if not splited_menu: continue
                    menu = splited_menu[0]
                    for k in range(1,len(splited_menu)):
                        menu += '/'
                        menu += splited_menu[k]
                    if tmpmenu[0] == '/' or flag == 1:
                        last = lunch.pop()
                        lunch.append(last + '/' + menu)
                        flag = 0
                    else: lunch.append(menu)
                    if tmpmenu[-1] == '/':
                        flag = 1
            for row in range(border3, border4):
                tmpmenu = worksheet.cell(row=row, column=day).value
                if tmpmenu:
                    splited_menu = re.compile('[가-힣]+').findall(tmpmenu)
                    if not splited_menu: continue
                    menu = splited_menu[0]
                    for k in range(1,len(splited_menu)):
                        menu += '/'
                        menu += splited_menu[k]
                    if tmpmenu[0] == '/' or flag == 1:
                        last = dinner.pop()
                        dinner.append(last + '/' + menu)
                        flag = 0
                    else: dinner.append(menu)
                    if tmpmenu[-1] == '/':
                        flag = 1
            if breakfast: menus.append(breakfast)
            if lunch: menus.append(lunch)
            if dinner: menus.append(dinner)
    menulist = []
    for i in range(len(menus)):
        menulist = menulist + menus[i]
    return menulist

allMenu = []
for i in range(1,9):
    allMenu = allMenu + eachMenus(data['Sheet'+str(i)])
allMenu = list(set(allMenu))
allMenu.sort()

f = open("menu.txt", 'w', encoding="UTF-8")
for menu in allMenu:
    f.write(menu + '\n')
f.close()